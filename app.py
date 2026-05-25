import csv
import html
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from main import scrape_subject


APP_TITLE = "Supplier IČO Lookup"
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

app = FastAPI(title=APP_TITLE)

DISCLAIMER = (
    "Tento výstup je podkladom pre interné posúdenie dodávateľa. "
    "Nejde o automatické odporúčanie ani právne, finančné alebo investičné stanovisko. "
    "Závery musia byť posúdené zodpovednou osobou v kontexte konkrétneho obchodného vzťahu."
)

SOURCE_LABELS = {
    "finstat": "FinStat",
    "orsr": "Obchodný register (ORSR)",
    "rpvs": "Register partnerov verejného sektora (RPVS)",
    "ruz": "Register účtovných závierok (RÚZ)",
}

SOURCE_SHORT_LABELS = {
    "finstat": "FinStat",
    "orsr": "ORSR",
    "rpvs": "RPVS",
    "ruz": "RÚZ",
}

SUMMARY_LABELS = {
    "ico": "IČO",
    "obchodne_meno": "Obchodné meno",
    "sidlo": "Sídlo",
    "pravna_forma": "Právna forma",
    "den_zapisu": "Deň zápisu",
    "zakladne_imanie": "Základné imanie",
    "sk_nace": "SK NACE",
    "kategoria_zamestnancov": "Kategória zamestnancov",
    "trzby_predaj_sluzieb": "Tržby z predaja služieb",
    "vynosy": "Výnosy",
    "zisk_strata": "Zisk / strata",
    "rpvs_pocet_kuv": "Počet KUV v RPVS",
    "orsr_pocet_statutarov": "Počet štatutárov v ORSR",
    "zdroj_orsr": "ORSR",
    "zdroj_rpvs": "RPVS",
    "zdroj_finstat": "FinStat",
    "zdroj_ruz": "RÚZ",
}

# ============================================================
# INPUT VALIDATION
# ============================================================

def normalize_ico(value: str) -> str:
    """
    Accepts:
    - 36785512
    - 36 785 512
    - 36-785-512

    Returns:
    - 36785512

    Raises:
    - ValueError if the value is not exactly 8 digits after cleanup.
    """
    ico = re.sub(r"\D", "", value or "")

    if not ico:
        raise ValueError("Zadajte IČO spoločnosti.")

    if len(ico) != 8:
        raise ValueError("Neplatný formát IČO.")

    return ico


def normalize_text(value: str) -> str:
    return " ".join(str(value).split())


# ============================================================
# RESULT NORMALIZATION
# ============================================================

def stringify(value: Any) -> str:
    """
    Converts nested dict/list values into readable strings for CSV/XLSX cells.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)

    return str(value)


def find_first_value(source: dict, candidate_keys: list[str]) -> str:
    """
    Finds the first matching value from a dictionary by candidate key fragments.

    Useful because FinStat/RÚZ labels can vary slightly.
    """
    if not isinstance(source, dict):
        return ""

    lowered_candidates = [x.lower() for x in candidate_keys]

    for key, value in source.items():
        key_lower = str(key).lower()

        if any(candidate in key_lower for candidate in lowered_candidates):
            return stringify(value)

    return ""


def flatten_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, dict):
        return " ".join(flatten_text(item) for item in value.values())

    if isinstance(value, list):
        return " ".join(flatten_text(item) for item in value)

    return str(value)


def parse_number(value: Any) -> float | None:
    text = stringify(value)
    if not text:
        return None

    normalized = (
        text.replace("\xa0", " ")
        .replace("EUR", "")
        .replace("€", "")
        .replace(",", ".")
    )
    match = re.search(r"-?\d[\d\s.]*", normalized)
    if not match:
        return None

    number = match.group(0).replace(" ", "")
    try:
        return float(number)
    except ValueError:
        return None


def source_error_message(source_key: str) -> str:
    return (
        f"Zdroj {SOURCE_LABELS.get(source_key, source_key)} je momentálne nedostupný. "
        "Údaje z tohto registra neboli zahrnuté do hodnotenia."
    )


def best_company_name(result: dict) -> str:
    orsr = result.get("orsr", {}) or {}
    rpvs = result.get("rpvs", {}) or {}
    finstat = result.get("finstat", {}) or {}
    ruz = result.get("ruz", {}) or {}

    finstat_basic = finstat.get("zakladne_udaje", {}) or {}
    rpvs_partner = rpvs.get("partner_verejneho_sektora", {}) or {}

    return (
        orsr.get("obchodne_meno", "")
        or ruz.get("nazov", "")
        or rpvs_partner.get("Obchodné meno", "")
        or find_first_value(finstat_basic, ["názov", "nazov", "obchodné meno", "obchodne meno"])
    )


def build_summary_row(result: dict) -> dict:
    """
    Creates one flat summary row for CSV/XLSX.
    Keeps this intentionally conservative: only uses values that are present.
    """
    orsr = result.get("orsr", {}) or {}
    rpvs = result.get("rpvs", {}) or {}
    finstat = result.get("finstat", {}) or {}
    ruz = result.get("ruz", {}) or {}

    finstat_basic = finstat.get("zakladne_udaje", {}) or {}
    finstat_financials = finstat.get("financne_ukazovatele", {}) or {}

    rpvs_partner = rpvs.get("partner_verejneho_sektora", {}) or {}
    rpvs_ubos = rpvs.get("konecni_uzivatelia_vyhod", []) or []

    return {
        "ico": result.get("ico", ""),
        "obchodne_meno": best_company_name(result),
        "sidlo": orsr.get("sidlo", "") or finstat_basic.get("Sídlo", ""),
        "pravna_forma": orsr.get("pravna_forma", ""),
        "den_zapisu": orsr.get("den_zapisu", ""),
        "zakladne_imanie": orsr.get("vyska_zakladneho_imania", ""),
        "sk_nace": finstat_basic.get("SK NACE", "") or ruz.get("SK NACE", ""),
        "kategoria_zamestnancov": finstat_basic.get("Kategória zamestnancov", ""),
        "trzby_predaj_sluzieb": find_first_value(
            finstat_financials,
            ["tržby z predaja služieb", "trzby z predaja sluzieb", "tržby"],
        ),
        "vynosy": find_first_value(
            finstat_financials,
            ["výnosy", "vynosy"],
        ),
        "zisk_strata": find_first_value(
            finstat_financials,
            ["zisk", "strata", "výsledok hospodárenia", "vysledok hospodarenia"],
        ),
        "rpvs_pocet_kuv": len(rpvs_ubos),
        "orsr_pocet_statutarov": len(orsr.get("statutarny_organ", []) or []),
        "zdroj_orsr": "áno" if bool(orsr) else "nie",
        "zdroj_rpvs": "áno" if bool(rpvs) else "nie",
        "zdroj_finstat": "áno" if bool(finstat) else "nie",
        "zdroj_ruz": "áno" if bool(ruz) else "nie",
    }


def build_risk_flags(result: dict) -> list[dict]:
    """
    Simple initial risk rules.
    Keep these transparent and easy to audit.
    """
    flags = []

    orsr = result.get("orsr", {}) or {}
    rpvs = result.get("rpvs", {}) or {}
    finstat = result.get("finstat", {}) or {}
    ruz = result.get("ruz", {}) or {}
    selected_sources = set(result.get("selected_sources") or SOURCE_LABELS.keys())

    if "orsr" in selected_sources and not orsr:
        flags.append({
            "severity": "Na overenie",
            "flag": "Údaje z ORSR nie sú dostupné",
            "detail": source_error_message("orsr"),
        })

    if "finstat" in selected_sources and not finstat:
        flags.append({
            "severity": "Na overenie",
            "flag": "Údaje z FinStat nie sú dostupné",
            "detail": source_error_message("finstat"),
        })

    if "ruz" in selected_sources and not ruz:
        flags.append({
            "severity": "Na overenie",
            "flag": "Údaje z RÚZ nie sú dostupné",
            "detail": source_error_message("ruz"),
        })

    if "rpvs" in selected_sources and not rpvs:
        flags.append({
            "severity": "Na overenie",
            "flag": "Údaje z RPVS nie sú dostupné",
            "detail": source_error_message("rpvs"),
        })

    selenium_error = result.get("selenium_error")
    if isinstance(selenium_error, dict):
        flags.append({
            "severity": "Na overenie",
            "flag": "Automatizované načítanie registrov zlyhalo",
            "detail": "Niektoré verejné registre sa nepodarilo načítať. Výsledok môže byť neúplný.",
        })

    if rpvs and not rpvs.get("konecni_uzivatelia_vyhod"):
        flags.append({
            "severity": "Na overenie",
            "flag": "RPVS nevrátil konečných užívateľov výhod",
            "detail": "RPVS bol dostupný, ale aplikácia nenašla záznamy KUV. Je vhodné overiť vlastnícke údaje manuálne.",
        })

    return flags


def get_registry_status(result: dict, summary_row: dict) -> dict:
    evidence = flatten_text({
        "summary": summary_row,
        "orsr": result.get("orsr", {}),
        "finstat": result.get("finstat", {}),
        "ruz": result.get("ruz", {}),
    }).lower()

    severe_patterns = [
        ("V konkurze", ["konkurz", "konkurze", "úpadok", "upadok", "insolv"]),
        ("V likvidácii", ["likvidácia", "likvidacii", "likvidácii"]),
        ("Vymazaná", ["vymazan", "zaniknut", "zrušen"]),
    ]

    for label, patterns in severe_patterns:
        if any(pattern in evidence for pattern in patterns):
            return {
                "label": label,
                "tone": "red",
                "reason": "V dostupných údajoch sa nachádza právne významný stav vyžadujúci zvýšenú pozornosť.",
            }

    if result.get("orsr") or result.get("finstat") or result.get("ruz"):
        return {
            "label": "Aktívna",
            "tone": "green",
            "reason": "Dostupné registre neobsahujú jednoznačný negatívny stav typu konkurz, likvidácia alebo výmaz.",
        }

    return {
        "label": "Neoverené",
        "tone": "gray",
        "reason": "Nie sú dostupné dostatočné údaje na určenie registračného stavu.",
    }


def build_registry_coverage(result: dict) -> list[dict]:
    selected_sources = set(result.get("selected_sources") or SOURCE_LABELS.keys())
    coverage = []

    for source_key, label in SOURCE_LABELS.items():
        source_data = result.get(source_key) or {}
        source_error = result.get(f"{source_key}_error")

        if source_key not in selected_sources:
            status = "Neoverené"
            detail = "Zdroj nebol vybraný pre toto vyhľadávanie."
        elif source_data and source_error:
            status = "Čiastočné"
            detail = "Zdroj vrátil niektoré údaje, ale načítanie nebolo úplné."
        elif source_data:
            status = "Získané"
            detail = "Údaje zo zdroja boli zahrnuté do podkladu."
        elif source_error:
            status = "Chyba"
            detail = source_error_message(source_key)
        else:
            status = "Nedostupné"
            detail = "Zdroj nevrátil použiteľné údaje."

        coverage.append({
            "source": source_key,
            "label": label,
            "status": status,
            "detail": detail,
        })

    return coverage


def confidence_item(label: str, level: str, reason: str) -> dict:
    return {"label": label, "level": level, "reason": reason}


def build_confidence_model(result: dict, coverage: list[dict], registry_status: dict) -> dict:
    selected = [item for item in coverage if item["status"] != "Neoverené"]
    obtained = [item for item in coverage if item["status"] in {"Získané", "Čiastočné"}]
    failed = [item for item in selected if item["status"] in {"Chyba", "Nedostupné"}]

    if not selected:
        completeness = confidence_item("Úplnosť údajov", "Neznáma", "Neboli vybrané žiadne zdroje.")
    elif len(obtained) == len(selected):
        completeness = confidence_item(
            "Úplnosť údajov",
            "Vysoká",
            f"Údaje boli získané zo všetkých {len(selected)} vybraných zdrojov.",
        )
    elif len(obtained) >= max(1, len(selected) // 2):
        completeness = confidence_item(
            "Úplnosť údajov",
            "Stredná",
            f"Údaje boli získané z {len(obtained)} z {len(selected)} vybraných zdrojov.",
        )
    else:
        completeness = confidence_item(
            "Úplnosť údajov",
            "Nízka",
            f"Údaje boli získané len z {len(obtained)} z {len(selected)} vybraných zdrojov.",
        )

    has_update_date = bool(find_first_value(result.get("orsr", {}) or {}, ["aktualizácie", "aktualizacie"]))
    has_financials = bool((result.get("finstat", {}) or {}).get("financne_ukazovatele"))
    if has_update_date and has_financials:
        freshness = confidence_item(
            "Aktuálnosť údajov",
            "Vysoká",
            "Dostupné sú finančné údaje aj údaj o aktualizácii aspoň jedného registra.",
        )
    elif has_financials or has_update_date:
        freshness = confidence_item(
            "Aktuálnosť údajov",
            "Stredná",
            "Niektoré údaje obsahujú časový kontext, ale nie pri všetkých zdrojoch je známa aktuálnosť.",
        )
    else:
        freshness = confidence_item(
            "Aktuálnosť údajov",
            "Neznáma",
            "Z dostupných údajov nie je jasné, kedy boli jednotlivé zdroje naposledy aktualizované.",
        )

    names = [
        result.get("orsr", {}).get("obchodne_meno"),
        result.get("ruz", {}).get("nazov"),
        (result.get("rpvs", {}).get("partner_verejneho_sektora", {}) or {}).get("Obchodné meno"),
    ]
    normalized_names = {
        normalize_text(str(name)).lower()
        for name in names
        if name
    }
    if len(normalized_names) <= 1 and obtained:
        consistency = confidence_item(
            "Konzistentnosť medzi registrami",
            "Vysoká",
            "Dostupné identifikačné údaje neukazujú zjavný rozpor medzi zdrojmi.",
        )
    elif len(normalized_names) == 2:
        consistency = confidence_item(
            "Konzistentnosť medzi registrami",
            "Stredná",
            "Niektoré názvy alebo identifikačné údaje sa líšia a môžu vyžadovať manuálne porovnanie.",
        )
    else:
        consistency = confidence_item(
            "Konzistentnosť medzi registrami",
            "Neznáma",
            "Nie je dostupný dostatok údajov na porovnanie medzi registrami.",
        )

    interpretation_level = "Stredná" if obtained and not failed else "Nízka"
    interpretation_reason = (
        "Interpretácia vychádza z verejne dostupných údajov a neslúži ako konečné rozhodnutie o dodávateľovi."
        if obtained
        else "Interpretácia je obmedzená, pretože sa nepodarilo získať dostatok údajov."
    )
    if registry_status["label"] in {"V konkurze", "V likvidácii", "Vymazaná"}:
        interpretation_level = "Vysoká"
        interpretation_reason = "Závažný právny stav je priamo viditeľný v dostupných údajoch."

    return {
        "completeness": completeness,
        "freshness": freshness,
        "consistency": consistency,
        "interpretation": confidence_item(
            "Spoľahlivosť interpretácie",
            interpretation_level,
            interpretation_reason,
        ),
    }


def build_key_observations(result: dict, summary_row: dict, coverage: list[dict], registry_status: dict) -> list[dict]:
    observations = []

    if registry_status["label"] in {"V konkurze", "V likvidácii", "Vymazaná"}:
        observations.append({
            "level": "Závažný signál",
            "title": f"Právny stav: {registry_status['label']}",
            "basis": registry_status["reason"],
            "caveat": "Tento signál je potrebné overiť v príslušnom registri pred akýmkoľvek obchodným rozhodnutím.",
        })
    elif registry_status["label"] == "Aktívna":
        observations.append({
            "level": "Informačné",
            "title": "Neboli identifikované jednoznačné signály konkurzu, likvidácie alebo výmazu",
            "basis": "Vyhodnotenie vychádza z dostupných údajov vo vybraných zdrojoch.",
            "caveat": "Absencia signálu v dostupných údajoch neznamená automatické schválenie dodávateľa.",
        })

    failed_sources = [item for item in coverage if item["status"] in {"Chyba", "Nedostupné"}]
    if failed_sources:
        observations.append({
            "level": "Na overenie",
            "title": "Niektoré registre neboli dostupné",
            "basis": "Nedostupné zdroje: " + ", ".join(item["label"] for item in failed_sources) + ".",
            "caveat": "Výsledok môže byť neúplný a vyžaduje doplňujúce manuálne overenie.",
        })

    employee_category = summary_row.get("kategoria_zamestnancov")
    revenue = parse_number(summary_row.get("vynosy") or summary_row.get("trzby_predaj_sluzieb"))
    if employee_category and revenue and revenue >= 1_000_000:
        observations.append({
            "level": "Na overenie",
            "title": "Kapacita spoločnosti môže vyžadovať doplňujúce overenie",
            "basis": f"Finančné údaje uvádzajú významnejšie výnosy a kategóriu zamestnancov: {employee_category}.",
            "caveat": "Tento pomer môže byť legitímny pri určitých obchodných modeloch, ale pri dodávateľskom posúdení môže byť vhodné overiť kapacity a subdodávateľské zabezpečenie.",
        })

    if summary_row.get("rpvs_pocet_kuv") == 0 and result.get("rpvs"):
        observations.append({
            "level": "Na overenie",
            "title": "RPVS neobsahuje parsované údaje o KUV",
            "basis": "RPVS bol dostupný, ale aplikácia nenašla konečných užívateľov výhod.",
            "caveat": "Vlastnícke údaje je vhodné overiť manuálne, najmä pri regulovaných alebo hodnotovo významných dodávkach.",
        })

    if not observations:
        observations.append({
            "level": "Informačné",
            "title": "Dostupné údaje sú obmedzené",
            "basis": "Vybrané zdroje nevrátili dostatok údajov na významnejšie pozorovania.",
            "caveat": "Pred interným rozhodnutím je vhodné doplniť overenie podľa významnosti dodávky.",
        })

    return observations


def build_verification_prompts(result: dict, coverage: list[dict], observations: list[dict]) -> list[str]:
    prompts = [
        "Je rozsah plánovanej dodávky primeraný dostupným údajom o kapacite spoločnosti?",
        "Sú finančné údaje dostatočne aktuálne pre interné hodnotenie?",
    ]

    if any(item["status"] in {"Chyba", "Nedostupné", "Neoverené"} for item in coverage):
        prompts.append("Je potrebné doplniť manuálne overenie registrov, ktoré neboli dostupné alebo neboli vybrané?")

    if result.get("rpvs") or "rpvs" in (result.get("selected_sources") or []):
        prompts.append("Sú vlastnícke a štatutárne údaje dostatočne overené?")

    if any(observation["level"] in {"Významný signál", "Závažný signál"} for observation in observations):
        prompts.append("Vyžaduje identifikovaný signál eskaláciu podľa interných pravidiel nákupu alebo compliance?")

    prompts.append("Je potrebné vyžiadať doplňujúce dokumenty od dodávateľa?")
    return prompts


def build_operational_overview(
    result: dict,
    summary_row: dict,
    coverage: list[dict],
    registry_status: dict,
    observations: list[dict],
) -> list[str]:
    company_name = summary_row.get("obchodne_meno") or "Spoločnosť"
    selected_count = sum(1 for item in coverage if item["status"] != "Neoverené")
    obtained_count = sum(1 for item in coverage if item["status"] in {"Získané", "Čiastočné"})

    paragraphs = [
        (
            f"{company_name} bola preverovaná podľa vybraných verejných zdrojov. "
            f"Údaje boli získané z {obtained_count} z {selected_count or 0} vybraných zdrojov."
        )
    ]

    if registry_status["label"] in {"V konkurze", "V likvidácii", "Vymazaná"}:
        paragraphs.append(
            f"Spoločnosť má podľa dostupných údajov stav „{registry_status['label']}“. "
            "Táto skutočnosť predstavuje závažný signál pre interné dodávateľské posúdenie."
        )
    elif registry_status["label"] == "Aktívna":
        paragraphs.append(
            "Dostupné údaje nenaznačujú jednoznačný stav konkurzu, likvidácie alebo výmazu. "
            "Tento výstup však nie je automatickým odporúčaním dodávateľa."
        )
    else:
        paragraphs.append(
            "Registračný stav sa nepodarilo spoľahlivo určiť z dostupných údajov. "
            "Pred rozhodnutím je vhodné doplniť manuálne overenie."
        )

    if any(observation["level"] in {"Na overenie", "Významný signál", "Závažný signál"} for observation in observations):
        paragraphs.append(
            "Niektoré zistenia vyžadujú ľudské posúdenie v kontexte plánovaného obchodného vzťahu, "
            "najmä pri významnej hodnote dodávky alebo regulovanom predmete plnenia."
        )

    paragraphs.append("Výstup slúži ako interný podklad pre posúdenie dodávateľa, nie ako finálny verdikt.")
    return paragraphs


def build_vendor_intelligence(result: dict, summary_row: dict) -> dict:
    retrieved_at = result.get("retrieved_at") or datetime.now().strftime("%d.%m.%Y %H:%M")
    result["retrieved_at"] = retrieved_at

    registry_status = get_registry_status(result, summary_row)
    coverage = build_registry_coverage(result)
    observations = build_key_observations(result, summary_row, coverage, registry_status)
    confidence = build_confidence_model(result, coverage, registry_status)
    prompts = build_verification_prompts(result, coverage, observations)
    overview = build_operational_overview(result, summary_row, coverage, registry_status, observations)

    return {
        "retrieved_at": retrieved_at,
        "registry_status": registry_status,
        "coverage": coverage,
        "confidence": confidence,
        "observations": observations,
        "verification_prompts": prompts,
        "overview": overview,
    }


# ============================================================
# OUTPUT WRITERS
# ============================================================

def output_base_name(ico: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{ico}_{timestamp}"


def write_json_output(result: dict, base_name: str) -> Path:
    path = OUTPUT_DIR / f"{base_name}.json"

    with path.open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return path


def write_csv_output(summary_row: dict, base_name: str) -> Path:
    path = OUTPUT_DIR / f"{base_name}_summary.csv"

    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(summary_row.keys()))
        writer.writeheader()
        writer.writerow(summary_row)

    return path


def autosize_sheet_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = stringify(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top")


def add_key_value_sheet(wb: Workbook, title: str, data: dict) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(["key", "value"])

    if isinstance(data, dict):
        for key, value in data.items():
            ws.append([str(key), stringify(value)])
    else:
        ws.append(["value", stringify(data)])

    style_header(ws)
    autosize_sheet_columns(ws)


def write_xlsx_output(
    result: dict,
    summary_row: dict,
    risk_flags: list[dict],
    base_name: str,
) -> Path:
    path = OUTPUT_DIR / f"{base_name}_report.xlsx"

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Súhrn"
    ws.append(list(summary_row.keys()))
    ws.append([stringify(value) for value in summary_row.values()])
    style_header(ws)
    autosize_sheet_columns(ws)

    # Signal sheet
    ws_flags = wb.create_sheet(title="Signály")
    ws_flags.append(["úroveň", "signál", "detail"])

    for flag in risk_flags:
        ws_flags.append([
            flag.get("severity", ""),
            flag.get("flag", ""),
            flag.get("detail", ""),
        ])

    style_header(ws_flags)
    autosize_sheet_columns(ws_flags)

    # Source-specific sheets
    add_key_value_sheet(wb, "ORSR", result.get("orsr", {}) or {})
    add_key_value_sheet(wb, "RPVS", result.get("rpvs", {}) or {})
    add_key_value_sheet(wb, "FinStat", result.get("finstat", {}) or {})
    add_key_value_sheet(wb, "RUZ", result.get("ruz", {}) or {})

    # Raw data sheet
    ws_raw = wb.create_sheet(title="Surové JSON")
    ws_raw.append(["raw_json"])
    ws_raw.append([json.dumps(result, ensure_ascii=False, indent=2)])
    style_header(ws_raw)
    ws_raw.column_dimensions["A"].width = 120
    ws_raw["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path


def generate_outputs(result: dict) -> dict:
    ico = result.get("ico", "unknown")
    base_name = output_base_name(ico)

    summary_row = build_summary_row(result)
    intelligence = build_vendor_intelligence(result, summary_row)
    risk_flags = build_risk_flags(result)

    json_path = write_json_output(result, base_name)
    csv_path = write_csv_output(summary_row, base_name)
    xlsx_path = write_xlsx_output(result, summary_row, risk_flags, base_name)

    return {
        "summary_row": summary_row,
        "intelligence": intelligence,
        "risk_flags": risk_flags,
        "json_file": json_path.name,
        "csv_file": csv_path.name,
        "xlsx_file": xlsx_path.name,
    }


# ============================================================
# HTML RENDERING
# ============================================================

def display_label(key: str) -> str:
    return SUMMARY_LABELS.get(key, key.replace("_", " ").strip().capitalize())


def render_value(value: Any) -> str:
    if value is None or value == "":
        return ""

    if isinstance(value, dict):
        if not value:
            return ""

        rows = "\n".join(
            f"<tr><th>{html.escape(display_label(str(key)))}</th><td>{render_value(item)}</td></tr>"
            for key, item in value.items()
        )
        return f"<table class=\"nested-table\">{rows}</table>"

    if isinstance(value, list):
        if not value:
            return ""

        items = "\n".join(f"<li>{render_value(item)}</li>" for item in value)
        return f"<ul>{items}</ul>"

    return html.escape(str(value))


def render_key_value_table(data: dict) -> str:
    rows = "\n".join(
        f"<tr><th>{html.escape(display_label(str(key)))}</th><td>{render_value(value)}</td></tr>"
        for key, value in data.items()
        if value not in (None, "", [], {})
    )

    if not rows:
        return "<p class=\"muted\">Bez dostupných údajov.</p>"

    return f"<table>{rows}</table>"


def render_source_sections(result: dict) -> str:
    selected_sources = result.get("selected_sources") or list(SOURCE_LABELS.keys())
    sections = []

    for source_key in selected_sources:
        label = SOURCE_LABELS.get(source_key, source_key)
        source_data = result.get(source_key) or {}
        source_error = result.get(f"{source_key}_error") or {}

        if source_data:
            content = render_key_value_table(source_data)
        elif isinstance(source_error, dict) and source_error:
            content = f"<p class=\"source-warning\">{html.escape(source_error_message(source_key))}</p>"
        else:
            content = "<p class=\"muted\">Zdroj bol vybraný, ale nevrátil údaje.</p>"

        sections.append(f"""
        <details class="source-section">
            <summary>{html.escape(label)}</summary>
            {content}
        </details>
        """)

    return "\n".join(sections)


def selected_source_copy(selected_sources: list[str]) -> str:
    labels = [SOURCE_SHORT_LABELS[source] for source in selected_sources if source in SOURCE_SHORT_LABELS]

    if not labels:
        return "Nebolo vybrané žiadne vyhľadávanie."

    if len(labels) == 1:
        return f"Vybraný zdroj: {labels[0]}."

    return "Vybrané zdroje: " + ", ".join(labels) + "."


def render_paragraphs(paragraphs: list[str]) -> str:
    return "\n".join(f"<p>{html.escape(paragraph)}</p>" for paragraph in paragraphs)


def render_confidence_cards(confidence: dict) -> str:
    cards = []
    for item in confidence.values():
        cards.append(f"""
        <div class="metric-card">
            <div class="metric-label">{html.escape(item["label"])}</div>
            <div class="metric-level">{html.escape(item["level"])}</div>
            <p>{html.escape(item["reason"])}</p>
        </div>
        """)
    return "<div class=\"card-grid\">" + "\n".join(cards) + "</div>"


def render_observation_cards(observations: list[dict]) -> str:
    return "\n".join(f"""
    <article class="stack-card">
        <div class="signal">{html.escape(item["level"])}</div>
        <h3>{html.escape(item["title"])}</h3>
        <p><strong>Základ:</strong> {html.escape(item["basis"])}</p>
        <p><strong>Poznámka:</strong> {html.escape(item["caveat"])}</p>
    </article>
    """ for item in observations)


def render_prompt_list(prompts: list[str]) -> str:
    return "<ul class=\"prompt-list\">" + "\n".join(
        f"<li>{html.escape(prompt)}</li>"
        for prompt in prompts
    ) + "</ul>"


def render_coverage_cards(coverage: list[dict]) -> str:
    return "<div class=\"card-grid\">" + "\n".join(f"""
    <div class="coverage-card">
        <div class="coverage-top">
            <strong>{html.escape(SOURCE_SHORT_LABELS.get(item["source"], item["label"]))}</strong>
            <span class="coverage-status">{html.escape(item["status"])}</span>
        </div>
        <p>{html.escape(item["detail"])}</p>
    </div>
    """ for item in coverage) + "</div>"


def render_download_section(outputs: dict) -> str:
    return f"""
    <section class="panel">
        <h2>Stiahnuť podklady</h2>
        <div class="downloads">
            <a href="/download/{html.escape(outputs["xlsx_file"])}">Dátový export XLSX</a>
            <a href="/download/{html.escape(outputs["csv_file"])}">Dátový export CSV</a>
            <a href="/download/{html.escape(outputs["json_file"])}">Dátový export JSON</a>
        </div>
        <p class="muted">Súhrnný report PDF je v backloge.</p>
    </section>
    """


def render_company_header(summary_row: dict, intelligence: dict) -> str:
    company_name = summary_row.get("obchodne_meno") or "Názov spoločnosti nebol získaný"
    legal_form = summary_row.get("pravna_forma") or "Neoverené"
    registry_status = intelligence["registry_status"]

    return f"""
    <section class="company-header">
        <div>
            <p class="eyebrow">Interný podklad pre posúdenie dodávateľa</p>
            <h1>{html.escape(company_name)}</h1>
            <div class="company-meta">
                <span>IČO: {html.escape(summary_row.get("ico", ""))}</span>
                <span>Právna forma: {html.escape(legal_form)}</span>
                <span>Dátum overenia: {html.escape(intelligence["retrieved_at"])}</span>
            </div>
        </div>
        <span class="status-badge status-{html.escape(registry_status["tone"])}">{html.escape(registry_status["label"])}</span>
    </section>
    """


def render_page(title: str, body: str) -> HTMLResponse:
    return HTMLResponse(f"""
    <!doctype html>
    <html lang="sk">
    <head>
        <meta charset="utf-8">
        <title>{html.escape(title)}</title>
        <style>
            body {{
                background: #f7f8fa;
                color: #111827;
                font-family: Arial, sans-serif;
                line-height: 1.45;
                margin: 0;
                padding: 0;
            }}
            main {{
                max-width: 1120px;
                margin: 0 auto;
                padding: 42px 24px 64px;
            }}
            h1, h2, h3 {{
                line-height: 1.18;
            }}
            h1 {{
                margin: 0 0 12px;
            }}
            h2 {{
                margin: 0 0 16px;
            }}
            input {{
                padding: 10px;
                width: min(100%, 360px);
                font-size: 16px;
            }}
            fieldset {{
                border: 1px solid #ddd;
                margin: 18px 0 0;
                padding: 14px 16px 16px;
                max-width: 720px;
            }}
            legend {{
                font-weight: 700;
                padding: 0 6px;
            }}
            label {{
                display: inline-flex;
                align-items: center;
                gap: 6px;
                margin: 8px 18px 0 0;
                color: #333;
            }}
            input[type="checkbox"] {{
                width: auto;
            }}
            button {{
                padding: 10px 16px;
                font-size: 16px;
                cursor: pointer;
            }}
            button:disabled {{
                cursor: wait;
                opacity: 0.7;
            }}
            .actions {{
                margin-top: 14px;
            }}
            .panel {{
                background: #fff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                margin-top: 20px;
                padding: 20px;
            }}
            .home-card {{
                background: #fff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                padding: 24px;
                max-width: 760px;
            }}
            .value-prop {{
                color: #4b5563;
                font-size: 18px;
                margin: 0 0 24px;
            }}
            .company-header {{
                align-items: flex-start;
                background: #fff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                display: flex;
                gap: 20px;
                justify-content: space-between;
                padding: 24px;
            }}
            .eyebrow {{
                color: #475569;
                font-size: 13px;
                font-weight: 700;
                letter-spacing: 0.04em;
                margin: 0 0 8px;
                text-transform: uppercase;
            }}
            .company-meta {{
                color: #4b5563;
                display: flex;
                flex-wrap: wrap;
                gap: 10px 18px;
            }}
            .status-badge {{
                border-radius: 999px;
                display: inline-block;
                font-weight: 700;
                padding: 8px 12px;
                white-space: nowrap;
            }}
            .status-green {{
                background: #dcfce7;
                color: #166534;
            }}
            .status-red {{
                background: #fee2e2;
                color: #991b1b;
            }}
            .status-gray {{
                background: #e5e7eb;
                color: #374151;
            }}
            .card-grid {{
                display: grid;
                gap: 14px;
                grid-template-columns: repeat(auto-fit, minmax(230px, 1fr));
            }}
            .metric-card, .coverage-card, .stack-card {{
                background: #fff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                padding: 16px;
            }}
            .metric-label {{
                color: #4b5563;
                font-size: 14px;
            }}
            .metric-level {{
                font-size: 22px;
                font-weight: 700;
                margin: 6px 0;
            }}
            .signal, .coverage-status {{
                background: #eef2ff;
                border-radius: 999px;
                color: #3730a3;
                display: inline-block;
                font-size: 13px;
                font-weight: 700;
                padding: 4px 8px;
            }}
            .coverage-top {{
                align-items: center;
                display: flex;
                gap: 10px;
                justify-content: space-between;
                min-height: 32px;
            }}
            .coverage-top strong {{
                font-size: 22px;
                line-height: 1;
            }}
            .coverage-card {{
                display: flex;
                flex-direction: column;
                min-height: 150px;
            }}
            .coverage-card p {{
                margin-top: 18px;
            }}
            .prompt-list {{
                margin: 0;
                padding-left: 22px;
            }}
            .prompt-list li {{
                margin: 8px 0;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-top: 20px;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                vertical-align: top;
                text-align: left;
            }}
            th {{
                background: #f2f2f2;
            }}
            pre {{
                background: #f7f7f7;
                padding: 16px;
                overflow-x: auto;
                border: 1px solid #ddd;
            }}
            .hint {{
                color: #666;
                margin-top: 8px;
            }}
            .muted {{
                color: #666;
            }}
            .status {{
                color: #4b5563;
                display: none;
                margin-top: 14px;
            }}
            .summary-note {{
                background: #eef6ff;
                border: 1px solid #bfdbfe;
                color: #1e3a8a;
                padding: 12px;
                margin: 16px 0;
            }}
            .downloads a {{
                background: #111827;
                border-radius: 6px;
                color: #fff;
                display: inline-block;
                margin-right: 12px;
                margin-bottom: 12px;
                padding: 10px 12px;
                text-decoration: none;
            }}
            .source-section {{
                margin-top: 32px;
            }}
            .source-section summary {{
                cursor: pointer;
                font-size: 20px;
                font-weight: 700;
                margin-bottom: 12px;
            }}
            .nested-table {{
                margin-top: 0;
            }}
            .source-warning {{
                background: #fff8e6;
                border: 1px solid #f2d28b;
                color: #7a4b00;
                padding: 12px;
            }}
            .recent-searches {{
                margin-top: 24px;
            }}
            .recent-item {{
                display: inline-block;
                margin: 0 10px 10px 0;
            }}
            .recent-item button {{
                background: #fff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                color: #111827;
                min-width: 220px;
                padding: 12px;
                text-align: left;
            }}
            .recent-item span {{
                color: #6b7280;
                display: block;
                font-size: 13px;
                margin-top: 4px;
            }}
            .error {{
                color: #9b1c1c;
                background: #fff0f0;
                padding: 12px;
                border: 1px solid #e0b4b4;
            }}
        </style>
        <script>
            function markSubmitting(form) {{
                const icoInput = form.querySelector("input[name='ico']");
                const button = form.querySelector("button[type='submit']");
                const status = document.getElementById("lookup-status");
                const selectedServices = Array.from(form.querySelectorAll("input[name='services']:checked"));

                if (!icoInput || !icoInput.value.replace(/\\D/g, "")) {{
                    if (status) {{
                        status.style.display = "block";
                        status.textContent = "Zadajte IČO spoločnosti.";
                    }}
                    if (icoInput) {{
                        icoInput.focus();
                    }}
                    return false;
                }}

                if (button) {{
                    button.disabled = true;
                    button.textContent = selectedServices.length === 1
                        ? "Vyhľadávam..."
                        : "Kontrolujem vybrané zdroje...";
                }}

                if (status) {{
                    status.style.display = "block";
                    status.textContent = "Čakajte, výsledok pripravujeme podľa vybraných zdrojov.";
                }}

                return true;
            }}

            const defaultServices = ["finstat", "orsr", "rpvs", "ruz"];

            function recentSearches() {{
                try {{
                    return JSON.parse(localStorage.getItem("recentIcoSearches") || "[]");
                }} catch (error) {{
                    return [];
                }}
            }}

            function normalizeServices(services) {{
                const allowed = new Set(defaultServices);
                const normalized = Array.isArray(services)
                    ? services.filter((service) => allowed.has(service))
                    : [];
                return normalized.length ? normalized : defaultServices;
            }}

            function saveRecentSearch(ico, name, services) {{
                const current = recentSearches().filter((item) => item.ico !== ico);
                current.unshift({{
                    ico,
                    name: name || "",
                    services: normalizeServices(services),
                    searchedAt: new Date().toISOString()
                }});
                localStorage.setItem("recentIcoSearches", JSON.stringify(current.slice(0, 5)));
            }}

            function escapeHtml(value) {{
                return String(value || "")
                    .replaceAll("&", "&amp;")
                    .replaceAll("<", "&lt;")
                    .replaceAll(">", "&gt;")
                    .replaceAll('"', "&quot;")
                    .replaceAll("'", "&#039;");
            }}

            function renderRecentSearches() {{
                const container = document.getElementById("recent-searches");
                if (!container) {{
                    return;
                }}

                const searches = recentSearches();
                if (!searches.length) {{
                    container.innerHTML = "<p class='muted'>Zatiaľ žiadne vyhľadávania.</p>";
                    return;
                }}

                container.innerHTML = searches.map((item) => {{
                    const safeIco = escapeHtml(item.ico);
                    const hasName = item.name && item.name !== "Neznáma spoločnosť";
                    const title = hasName
                        ? escapeHtml(item.name)
                        : `IČO: ${{safeIco}}`;
                    const servicesInputs = normalizeServices(item.services)
                        .map((service) => `<input type="hidden" name="services" value="${{escapeHtml(service)}}">`)
                        .join("");

                    return `
                    <form method="post" action="/lookup" class="recent-item">
                        <input type="hidden" name="ico" value="${{safeIco}}">
                        ${{servicesInputs}}
                        <button type="submit">
                            <strong>${{title}}</strong>
                            <span>${{hasName ? `IČO: ${{safeIco}}` : "Zopakovať vyhľadávanie"}}</span>
                        </button>
                    </form>
                `; }}).join("");
            }}

            document.addEventListener("DOMContentLoaded", renderRecentSearches);
        </script>
    </head>
    <body>
        <main>{body}</main>
    </body>
    </html>
    """)


@app.get("/", response_class=HTMLResponse)
def home():
    return render_page(
        APP_TITLE,
        """
        <section class="home-card">
        <h1>Overenie dodávateľa podľa IČO</h1>
        <p class="value-prop">
            Získajte prehľad dostupných údajov z verejných registrov a podklady pre interné posúdenie dodávateľa.
        </p>

        <form method="post" action="/lookup" onsubmit="return markSubmitting(this)" novalidate>
            <input
                name="ico"
                placeholder="Zadajte IČO spoločnosti"
                autofocus
            >

            <fieldset>
                <legend>Zdroje na overenie</legend>
                <label>
                    <input type="checkbox" name="services" value="finstat" checked>
                    FinStat
                </label>
                <label>
                    <input type="checkbox" name="services" value="orsr" checked>
                    ORSR
                </label>
                <label>
                    <input type="checkbox" name="services" value="rpvs" checked>
                    RPVS
                </label>
                <label>
                    <input type="checkbox" name="services" value="ruz" checked>
                    RÚZ
                </label>
            </fieldset>

            <div class="actions">
                <button type="submit">Overiť dodávateľa</button>
            </div>
        </form>

        <div class="hint">
            Vyberte jeden alebo viac zdrojov. FinStat býva najrýchlejší; ORSR, RPVS a RÚZ môžu trvať dlhšie.
        </div>
        <div id="lookup-status" class="status"></div>
        </section>

        <section class="recent-searches">
            <h2>Posledné vyhľadávania</h2>
            <div id="recent-searches"></div>
        </section>
        """,
    )


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/lookup", response_class=HTMLResponse)
def lookup(ico: str = Form(""), services: list[str] = Form(["finstat", "orsr", "rpvs", "ruz"])):
    try:
        normalized_ico = normalize_ico(ico)
        selected_services = [
            service
            for service in services
            if service in SOURCE_LABELS
        ] or ["finstat"]

        result = scrape_subject(
            normalized_ico,
            selected_sources=selected_services,
        )
        outputs = generate_outputs(result)

        summary_row = outputs["summary_row"]
        intelligence = outputs["intelligence"]
        risk_flags = outputs["risk_flags"]

        source_sections_html = render_source_sections(result)
        incomplete_warning = (
            "<div class=\"summary-note\">Niektoré registre sú momentálne nedostupné. Výsledok môže byť neúplný.</div>"
            if any(item["status"] in {"Chyba", "Nedostupné"} for item in intelligence["coverage"])
            else ""
        )
        company_name_for_recent = summary_row.get("obchodne_meno") or ""
        recent_ico_json = json.dumps(normalized_ico, ensure_ascii=False)
        recent_name_json = json.dumps(company_name_for_recent, ensure_ascii=False)
        recent_services_json = json.dumps(selected_services, ensure_ascii=False)

        body = f"""
        <a href="/">← Nové vyhľadanie</a>

        {render_company_header(summary_row, intelligence)}
        {incomplete_warning}

        <section class="panel">
            <h2>Prevádzkový prehľad</h2>
            {render_paragraphs(intelligence["overview"])}
            <p class="muted">{html.escape(DISCLAIMER)}</p>
        </section>

        <section class="panel">
            <h2>Kvalita dostupných údajov</h2>
            {render_confidence_cards(intelligence["confidence"])}
        </section>

        <section class="panel">
            <h2>Kľúčové pozorovania</h2>
            {render_observation_cards(intelligence["observations"])}
        </section>

        <section class="panel">
            <h2>Pokrytie registrov</h2>
            {render_coverage_cards(intelligence["coverage"])}
        </section>

        {render_download_section(outputs)}

        <section class="panel">
            <h2>Detailné údaje zo zdrojov</h2>
            <p class="muted">Surové údaje sú dostupné v exportoch. Nižšie sú voliteľné detailné výpisy z vybraných zdrojov.</p>
            {source_sections_html}
        </section>

        <script>
            saveRecentSearch({recent_ico_json}, {recent_name_json}, {recent_services_json});
        </script>
        """

        return render_page(f"Výsledok {normalized_ico}", body)

    except ValueError as e:
        body = f"""
        <a href="/">← Späť</a>
        <h1>Vyhľadávanie sa nedá spustiť</h1>
        <div class="error">{html.escape(str(e))}</div>
        """

        return render_page("Chyba", body)

    except Exception:
        body = f"""
        <a href="/">← Späť</a>
        <h1>Vyhľadávanie zlyhalo</h1>
        <div class="error">Vyhľadávanie zlyhalo. Skúste to znova alebo overte zadané IČO.</div>
        """

        return render_page("Chyba", body)


@app.get("/download/{filename}")
def download_file(filename: str):
    """
    Safe download endpoint.
    Prevents path traversal by stripping directory parts from the filename.
    """
    safe_name = Path(filename).name
    path = OUTPUT_DIR / safe_name

    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Súbor sa nenašiel")

    suffix = path.suffix.lower()

    if suffix == ".json":
        media_type = "application/json"
    elif suffix == ".csv":
        media_type = "text/csv"
    elif suffix == ".xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        media_type = "application/octet-stream"

    return FileResponse(
        path=path,
        filename=safe_name,
        media_type=media_type,
    )
